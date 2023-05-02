import datetime
import openpyxl
from Utils.User import UserData
import json
import os

class OrgData:
    def __init__(self, org: str):
        checkfile = os.path.exists(f"Data/{org}")
        if(checkfile == False):
            os.mkdir(f"Data/{org}")
        else:
            pass

        today = datetime.date.today()
        start = today - datetime.timedelta(days=today.weekday())
        end = start + datetime.timedelta(days=6)
        file_name = f"{start.strftime('%d')}-{end.strftime('%d-%m-%Y')}.xlsx"
        with open("config.json", "r") as file:
            data = json.load(file)
        file.close()
        if(today.strftime("%d/%m/%Y") > data['week']['end']):
            self.wb = openpyxl.load_workbook("example.xlsx")
            with open("config.json", "w") as file:
                data['week']['start'] = start.strftime("%d/%m/%Y")
                data['week']['end'] = end.strftime("%d/%m/%Y")
                json.dump(data, file, indent=4)
            self.ws = self.wb.active
            self.dirs = f"Data/{org}/{file_name}"
        else:
            try:
                self.wb = openpyxl.load_workbook(f"Data/{org}/{file_name}")
                self.ws = self.wb.active
            except FileNotFoundError:
                self.wb = openpyxl.load_workbook("example.xlsx")
                self.ws = self.wb.active
                self.ws.cell(row=1, column=4, value=f"Danh sách học sinh vi phạm từ ngày {start.strftime('%d/%m/%Y')} đến ngày {end.strftime('%d/%m/%Y')} đơn vị({org})")
        self.user = UserData()
        self.org = org
        self.file_name = f"Data/{org}/{file_name}"
        
    def InsertToData(self, id: int,reason: str, note: str = ""):
        try:
            day = datetime.datetime.now().strftime("%d/%m/%Y-%H:%M:%S")
            name = self.user.GetUserFromID(id)["name"]
            clas = self.user.GetUserFromID(id)["class"]
            max_row = self.ws.max_row
            row_num = max_row + 1
            if(reason=="chậm"):
                self.ws.cell(row=row_num, column=1, value=row_num - 4)
                self.ws.cell(row=row_num, column=2, value=day)
                self.ws.cell(row=row_num, column=3, value=id)
                self.ws.cell(row=row_num, column=4, value=name)
                self.ws.cell(row=row_num, column=5, value=clas)
                self.ws.cell(row=row_num, column=6, value="x")
                self.ws.cell(row=row_num, column=12, value=note)
            elif(reason=="bỏ giờ"):
                self.ws.cell(row=row_num, column=1, value=row_num - 4)
                self.ws.cell(row=row_num, column=2, value=day)
                self.ws.cell(row=row_num, column=3, value=id)
                self.ws.cell(row=row_num, column=4, value=name)
                self.ws.cell(row=row_num, column=5, value=clas)
                self.ws.cell(row=row_num, column=7, value="x")
                self.ws.cell(row=row_num, column=12, value=note)
            elif(reason=="phù hiệu"):
                self.ws.cell(row=row_num, column=1, value=row_num - 4)
                self.ws.cell(row=row_num, column=2, value=day)
                self.ws.cell(row=row_num, column=3, value=id)
                self.ws.cell(row=row_num, column=4, value=name)
                self.ws.cell(row=row_num, column=5, value=clas)
                self.ws.cell(row=row_num, column=8, value="x")
                self.ws.cell(row=row_num, column=12, value=note)
            elif(reason=="đồng phục"):
                self.ws.cell(row=row_num, column=1, value=row_num - 4)
                self.ws.cell(row=row_num, column=2, value=day)
                self.ws.cell(row=row_num, column=3, value=id)
                self.ws.cell(row=row_num, column=4, value=name)
                self.ws.cell(row=row_num, column=5, value=clas)
                self.ws.cell(row=row_num, column=9, value="x")
                self.ws.cell(row=row_num, column=12, value=note)
            elif(reason=="điện thoại"):
                self.ws.cell(row=row_num, column=1, value=row_num - 4)
                self.ws.cell(row=row_num, column=2, value=day)
                self.ws.cell(row=row_num, column=3, value=id)
                self.ws.cell(row=row_num, column=4, value=name)
                self.ws.cell(row=row_num, column=5, value=clas)
                self.ws.cell(row=row_num, column=10, value="x")
                self.ws.cell(row=row_num, column=12, value=note)
            elif(reason=="khác"):
                self.ws.cell(row=row_num, column=1, value=row_num - 4)
                self.ws.cell(row=row_num, column=2, value=day)
                self.ws.cell(row=row_num, column=3, value=id)
                self.ws.cell(row=row_num, column=4, value=name)
                self.ws.cell(row=row_num, column=5, value=clas)
                self.ws.cell(row=row_num, column=11, value="x")
                self.ws.cell(row=row_num, column=12, value=note)
            self.wb.save(self.file_name)
        except Exception as error:
            print(error)
            
class SheetData():
    def __init__(self,org: str,clas:str):
        if(os.path.exists(f"Check/{org}")==False):
            os.mkdir(f"Check/{org}")
        else:
            try:
                self.wb = openpyxl.load_workbook(f"Check/{org}/{clas}.xlsx")
            except FileNotFoundError:
                self.wb = openpyxl.load_workbook("example_check.xlsx")
                self.wb.save(f"Check/{org}/{clas}.xlsx")
                self.wb = openpyxl.load_workbook(f"Check/{org}/{clas}.xlsx")
    
        self.org = org
        self.clas = clas
        self.file_name = f"Check/{org}/{clas}.xlsx"
        self.ws = self.wb.active
        self.user = UserData()

    def MakeNewData(self,name_column:str):
        temp = self.ws.max_column
        self.ws.cell(row=3,column=temp+1,value=name_column)
        self.wb.save(self.file_name)

    def InsertToData(self,id,password,name_column:str,note:str):
        name = self.user.GetUser(id,password)["name"]
        clas = self.user.GetUser(id,password)["class"]
        day = datetime.datetime.now().strftime("%d/%m/%Y - %H:%M:%S")
        temp_column = 0
        temp_row = 0
        for i in range(1,self.ws.max_column+1):
            if(self.ws.cell(row=3,column=i).value==name_column):
                temp_column = i
                break
        
        for i in range(1,self.ws.max_row+1):
            if(self.ws.cell(row=i,column=3).value==name):
                temp_row = i
                break
        
        if(temp_row==0):
            self.ws.cell(row=self.ws.max_row+1,column=1,value=self.ws.max_row-2)
            self.ws.cell(row=self.ws.max_row,column=2,value=day)
            self.ws.cell(row=self.ws.max_row,column=3,value=id)
            self.ws.cell(row=self.ws.max_row,column=4,value=name)
            self.ws.cell(row=self.ws.max_row,column=5,value=clas)
            self.ws.cell(row=self.ws.max_row,column=6,value=note)
            if(temp_column!=0):
                self.ws.cell(row=self.ws.max_row,column=temp_column,value="x")
            else:
                print("Không tìm thấy dữ liệu")
        elif(temp_column!=0 and temp_row!=0):
            self.ws.cell(row=temp_row,column=temp_column,value="x")
            self.ws.cell(row=temp_row,column=12,value=note)
        else:
            print("Không tìm thấy dữ liệu")                 
        self.wb.save(self.file_name)

class DataBus():
    def __init__(self,id_bus,org:str):
        self.id_bus = id_bus
        self.user = UserData()
        if(os.path.exists(path=f"Bus/{org}")==False):
            os.mkdir(path=f"Bus/{org}")
        try:
            self.wb = openpyxl.load_workbook(filename=f"Bus/{org}/{id_bus}.xlsx")
        except FileNotFoundError:
            self.wb = openpyxl.load_workbook(filename="example_bus.xlsx")
            self.wb.save(filename=f"Bus/{org}/{id_bus}.xlsx")
            self.wb = openpyxl.load_workbook(filename=f"Bus/{org}/{id_bus}.xlsx")

        self.file_name = f"Bus/{org}/{id_bus}.xlsx"

    def InsertToData(self,user_id,note:str):

        name = self.user.GetUserFromID(user_id)["name"]
        clas = self.user.GetUserFromID(user_id)["class"]
        day = datetime.datetime.now().strftime("%d-%m-%Y")
        time = datetime.datetime.now().strftime("%H:%M:%S")
        if(day not in self.wb.sheetnames):
            self.wb.create_sheet(day)
            self.wb.save(filename=self.file_name)
            self.ws = self.wb[day]
            self.ws.cell(row=1,column=1,value="STT")
            self.ws.cell(row=1,column=2,value="Thời gian")
            self.ws.cell(row=1,column=3,value="Mã định danh")
            self.ws.cell(row=1,column=4,value="Họ và tên")
            self.ws.cell(row=1,column=5,value="Lớp")
            self.ws.cell(row=1,column=6,value="Ghi chú")
            self.wb.save(filename=self.file_name)
        else:
            pass
        
        self.ws = self.wb[day]
        max_row = self.ws.max_row
        self.ws.cell(row=max_row+1,column=1,value=max_row)
        self.ws.cell(row=max_row+1,column=2,value=time)
        self.ws.cell(row=max_row+1,column=3,value=user_id)
        self.ws.cell(row=max_row+1,column=4,value=name)
        self.ws.cell(row=max_row+1,column=5,value=clas)
        self.ws.cell(row=max_row+1,column=6,value=note)
        self.wb.save(filename=self.file_name)


    

        




